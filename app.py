import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import io
from datetime import datetime

# ページ設定
st.set_page_config(page_title="湯瀬ホテル 在庫表 自動更新ツール", page_icon="♨️", layout="centered")

# タイトル
st.title("♨️ 湯瀬ホテル 在庫表 自動更新ツール")

# ファイルアップロード
uploaded_file = st.file_uploader(label="「年度在庫速度表 湯瀬」のExcelファイルをアップロードしてください", type=['xlsx'])

if uploaded_file is not None:
    try:
        # Excelファイルの読み込み
        wb = load_workbook(uploaded_file)
        
        # 必要なシートの存在確認
        required_sheets = ["当日", "前日", "特室在庫"]
        missing_sheets = [sheet for sheet in required_sheets if sheet not in wb.sheetnames]
        
        if missing_sheets:
            st.error(f"エラー: 必要なシートが見つかりません。不足シート: {', '.join(missing_sheets)}")
        else:
            with st.spinner("在庫表を更新しています..."):
                ws_today = wb["当日"]
                ws_yesterday = wb["前日"]
                ws_inventory = wb["特室在庫"]

                # 各客室の総室数定義（このリストが対象客室も兼ねます）
                room_capacities = {
                    "11F露ﾂｲﾝ": 3,
                    "11F露ｽｲｰﾄ": 4,
                    "4F露和": 4,
                    "803露和洋": 1,
                    "801半露洋洋": 1,
                    "802展望和洋": 1,
                    "157洋洋": 1,
                    "155洋": 1,
                    "7F和和": 2,
                    "7F和洋": 2
                }
                target_rooms = list(room_capacities.keys())

                # 塗りつぶしの色設定
                fill_blue = PatternFill(patternType='solid', fgColor='9BC2E6')  # 売れた時 (青)
                fill_red = PatternFill(patternType='solid', fgColor='F4B084')   # キャンセル時 (赤)
                fill_none = PatternFill(fill_type=None)                         # 色なし

                # データを抽出する関数
                def extract_data(sheet):
                    data = {}
                    dates = [cell.value for cell in sheet[1]]
                    for row in sheet.iter_rows(min_row=2):
                        room_name = row[0].value
                        if room_name in target_rooms:
                            data[room_name] = {}
                            for idx, cell in enumerate(row):
                                if idx == 0: continue
                                date_key = dates[idx]
                                if date_key is not None:
                                    date_str = str(date_key).split(' ')[0]
                                    data[room_name][date_str] = cell.value if cell.value is not None else 0
                    return dates, data

                # 当日と前日の販売数データを取得
                dates_today, data_today = extract_data(ws_today)
                dates_yday, data_yday = extract_data(ws_yesterday)

                # 特室在庫シートの更新
                inventory_dates = [cell.value for cell in ws_inventory[1]]
                
                for row in ws_inventory.iter_rows(min_row=2):
                    room_name = row[0].value
                    if room_name in target_rooms:
                        capacity = room_capacities[room_name]
                        
                        for idx, cell in enumerate(row):
                            if idx == 0: continue
                            
                            date_key = inventory_dates[idx]
                            if date_key is None: continue
                            
                            date_str = str(date_key).split(' ')[0]

                            # 当日と前日の「販売数」
                            sales_today = data_today.get(room_name, {}).get(date_str)
                            sales_yday = data_yday.get(room_name, {}).get(date_str)

                            if sales_today is not None:
                                # 値の転記（在庫数 = 総室数 - 当日の販売数）
                                inventory_count = capacity - int(sales_today)
                                
                                # ★修正ポイント：完売（0以下）の場合は空白（None）にする
                                if inventory_count <= 0:
                                    cell.value = None
                                else:
                                    cell.value = inventory_count
                                
                                # 色の判定（販売数の増減で判定）
                                if sales_yday is not None:
                                    if sales_today > sales_yday:
                                        # 販売数が増えた＝売れた（青）
                                        cell.fill = fill_blue
                                    elif sales_today < sales_yday:
                                        # 販売数が減った＝キャンセル（赤）
                                        cell.fill = fill_red
                                    else:
                                        cell.fill = fill_none

                # ダウンロード用のファイル名を作成 (yyyy.mm.dd 2025年度在庫速度表 湯瀬.xlsx)
                today_str = datetime.now().strftime("%Y.%m.%d")
                output_filename = f"{today_str} 2025年度在庫速度表 湯瀬.xlsx"

                # Excelファイルをメモリ上に保存
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

            st.success("処理が完了しました！以下のボタンからファイルをダウンロードしてください。")

            # ダウンロードボタン
            st.download_button(
                label="📥 更新されたファイルをダウンロード",
                data=output,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"エラーが発生しました: {e}")

# ------------------------------------
# 使用方法セクション（アップロードエリアより下部に配置）
# ------------------------------------
st.markdown("---")  # 区切り線
with st.expander("使用方法", expanded=True):
    st.markdown("""
    湯瀬ホテルの温泉付き客室の在庫の増減表を自動で作成するツールです。  
    正確なデータに基づいた在庫増減の確認ができます。

    **① 日々の在庫速度表を作成する** シェアフォルダ内でいつも通り、速度表の作成を開始します。

    支配人君から「日別部屋タイプ集計」をダウンロードし、「当日」にコピペします。  
    在庫は「前日」,「当日」シートの増減を基に計算します。  
    「特室在庫」の不要な日程を削除してください。（前日分など）  
    ⇒ 下準備はこれで完成です。

    **② 在庫表をドロップしてください** 上のアップロード枠に、下準備が終わった在庫表をドロップしてください。  
    在庫の増減を反映した新しいエクセルファイルをダウンロードできます。  
    その後、再度また シェアフォルダ にファイルを戻し、上書きしてください。

    **機能とルール** 「特室在庫」シートに最新の在庫数（総室数 - 販売数）が転記され、販売数の増減に応じてセルに色が付きます。  
    🔵 青色: 売れた（在庫が減少）  
    🔴 赤色: キャンセルが出た（在庫が増加）  
    ※ 完売（在庫0）の場合はセルが空白になります。
    """)
